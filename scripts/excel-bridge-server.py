import os
import sys
import json
import time
import datetime
import subprocess
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
PORT = 8787
API_KEY = "sf10-bridge-2024"
TEMPLATE_NAME = 'PLACEHOLDER(ALL).xlsx'

# Robust path detection for PyInstaller
if hasattr(sys, '_MEIPASS'):
    EXE_DIR = os.path.dirname(sys.executable)
else:
    EXE_DIR = os.path.dirname(os.path.abspath(__file__))

PATHS_TO_CHECK = [
    os.path.join(EXE_DIR, TEMPLATE_NAME),
    os.path.join(EXE_DIR, 'release', TEMPLATE_NAME),
    os.path.join(os.getcwd(), 'release', TEMPLATE_NAME),
    os.path.join(os.getcwd(), TEMPLATE_NAME)
]

TEMPLATE_PATH = TEMPLATE_NAME
for p in PATHS_TO_CHECK:
    if os.path.exists(p):
        TEMPLATE_PATH = p
        break

OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "form137-exports")
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR, exist_ok=True)

LOG_FILE = os.path.join(EXE_DIR, "excel-bridge-python.log")
def log_it(msg):
    try:
        with open(LOG_FILE, "a") as f:
            ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{ts}] {msg}\n")
    except:
        pass

# --- HELPER: SAFE WRITE ---
def safe_write(ws, coord, write_val):
    """Write a value to an Excel cell. Use None to clear the cell."""
    coord = str(coord).upper()
    try:
        ws[coord] = write_val
    except AttributeError:
        # Handle merged cells
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = write_val
                return
        log_it(f"Failed write: {coord}")

def to_num(val):
    """Convert grade string to int for Excel formulas. Returns None if empty (truly clears cell)."""
    if val is None or val == '':
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        try:
            return float(val)
        except (ValueError, TypeError):
            return val

def empty_to_none(val):
    """Convert empty strings to None so Excel cells are truly cleared."""
    if val is None or val == '':
        return None
    return val

def write_subject_row(ws, r, d):
    """Write a single subject row. Uses None to truly clear unused cells."""
    safe_write(ws, f'A{r}', empty_to_none(d.get('type', '')))
    safe_write(ws, f'I{r}', empty_to_none(d.get('subject', '')))
    safe_write(ws, f'AT{r}', to_num(d.get('q1', '')))
    safe_write(ws, f'AY{r}', to_num(d.get('q2', '')))
    safe_write(ws, f'BD{r}', to_num(d.get('final', '')))
    safe_write(ws, f'BI{r}', empty_to_none(d.get('action', '')))

def write_remedial_row(ws, r, rd):
    """Write a single remedial subject row. Uses None to truly clear unused cells."""
    safe_write(ws, f'A{r}', empty_to_none(rd.get('type', '')))
    safe_write(ws, f'I{r}', empty_to_none(rd.get('subject', '')))
    safe_write(ws, f'AT{r}', to_num(rd.get('semGrade', '')))
    safe_write(ws, f'AY{r}', to_num(rd.get('remedialMark', '')))
    safe_write(ws, f'BD{r}', to_num(rd.get('recomputedGrade', '')))
    safe_write(ws, f'BI{r}', empty_to_none(rd.get('action', '')))

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, allow_headers=["Content-Type", "X-Bridge-Key", "ngrok-skip-browser-warning"])

@app.route('/health', methods=['GET'])
def health():
    return jsonify({"success": True, "status": "Running", "template": TEMPLATE_PATH})

@app.route('/open-excel', methods=['POST'])
def open_excel():
    if API_KEY and request.headers.get('x-bridge-key') != API_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    try:
        body = request.get_json(); data = body.get('data', {}); info = data.get('info', {}); eligibility = data.get('eligibility', {}); cert = data.get('certification', {})
        return_file = body.get('returnFile', False)
        log_it(f"Exporting: {info.get('lname', 'Unknown')}")
        if not os.path.exists(TEMPLATE_PATH):
            log_it(f"ERR: Template missing at {TEMPLATE_PATH}")
            return jsonify({"success": False, "error": "Template missing"}), 500
        wb = openpyxl.load_workbook(TEMPLATE_PATH); sn = [s.upper() for s in wb.sheetnames]
        ws_front = wb[wb.sheetnames[sn.index('FRONT')]] if 'FRONT' in sn else wb.active
        ws_back = wb[wb.sheetnames[sn.index('BACK')]] if 'BACK' in sn else (wb[wb.sheetnames[1]] if len(wb.sheetnames)>1 else wb.active)

        # --- STUDENT INFO (FRONT) ---
        safe_write(ws_front, 'F8', info.get('lname', '')); safe_write(ws_front, 'Y8', info.get('fname', '')); safe_write(ws_front, 'AZ8', info.get('mname', ''))
        safe_write(ws_front, 'C9', info.get('lrn', '')); safe_write(ws_front, 'AN9', info.get('sex', '')); safe_write(ws_front, 'AA9', info.get('birthdate', ''))
        safe_write(ws_front, 'BH9', info.get('admissionDate', ''))
        
        # --- ELIGIBILITY (FRONT) ---
        safe_write(ws_front, 'Z14', eligibility.get('schoolName', '')); safe_write(ws_front, 'AW14', eligibility.get('schoolAddress', ''))
        if eligibility.get('hsCompleter'): safe_write(ws_front, 'A13', 'X')
        safe_write(ws_front, 'N13', eligibility.get('hsGenAve', ''))
        if eligibility.get('jhsCompleter'): safe_write(ws_front, 'S13', 'X')
        safe_write(ws_front, 'AH13', eligibility.get('jhsGenAve', '')); safe_write(ws_front, 'P14', eligibility.get('gradDate', ''))
        if eligibility.get('pept'): safe_write(ws_front, 'A16', 'X')
        safe_write(ws_front, 'K16', eligibility.get('peptRating', ''))
        if eligibility.get('als'): safe_write(ws_front, 'S16', 'X')
        safe_write(ws_front, 'AC16', eligibility.get('alsRating', ''))
        if eligibility.get('others'): safe_write(ws_front, 'AH16', 'X')
        safe_write(ws_front, 'AP16', eligibility.get('othersSpec', '')); safe_write(ws_front, 'P17', eligibility.get('examDate', ''))
        safe_write(ws_front, 'AN17', eligibility.get('clcName', ''))

        # ==========================================================
        # GRADE 11 (FRONT PAGE)
        # ==========================================================
        
        # --- SEMESTER 1 (G11 - SEM 1) ---
        s1 = data.get('semester1', {})
        safe_write(ws_front, 'E23', s1.get('school', '')); safe_write(ws_front, 'AF23', s1.get('schoolId', ''))
        safe_write(ws_front, 'AS23', s1.get('gradeLevel', '')); safe_write(ws_front, 'BA23', s1.get('sy', ''))
        safe_write(ws_front, 'BK23', s1.get('sem', '')); safe_write(ws_front, 'G25', s1.get('trackStrand', ''))
        safe_write(ws_front, 'AS25', s1.get('section', ''))

        sub1 = s1.get('subjects', [])
        for i, r in enumerate(range(31, 43)): # 11 rows
            d = sub1[i] if i < len(sub1) else {}
            write_subject_row(ws_front, r, d if d else {})

        # Gen Ave: write our calculated value (template formula is unreliable)
        safe_write(ws_front, 'BD43', to_num(s1.get('genAve', '')))
        safe_write(ws_front, 'E45', s1.get('remarks', ''))
        # --- Signatory (Sem 1) --- PLACEHOLDER CELLS, calibrate as needed
        safe_write(ws_front, 'A49', s1.get('adviserName', ''))
        safe_write(ws_front, 'Y49', s1.get('certName', ''))
        safe_write(ws_front, 'AZ49', s1.get('dateChecked', ''))

        r1 = s1.get('remedial', {})
        if r1:
            safe_write(ws_front, 'S52', r1.get('from', '')); safe_write(ws_front, 'AC52', r1.get('to', ''))
            safe_write(ws_front, 'AL52', r1.get('school', '')); safe_write(ws_front, 'BK52', r1.get('schoolId', ''))
            rem_subs1 = r1.get('subjects', [])
            for i, r in enumerate(range(58, 63)): # 5 remedial rows
                rd = rem_subs1[i] if i < len(rem_subs1) else {}
                write_remedial_row(ws_front, r, rd if rd else {})
            # Remedial teacher/signature — PLACEHOLDER CELLS, calibrate as needed
            safe_write(ws_front, 'J63', r1.get('teacherName', ''))
            safe_write(ws_front, 'AY63', r1.get('signature', ''))

        # --- SEMESTER 2 (G11 - SEM 2) ---
        s2 = data.get('semester2', {})
        safe_write(ws_front, 'E66', s2.get('school', '')); safe_write(ws_front, 'AF66', s2.get('schoolId', ''))
        safe_write(ws_front, 'AS66', s2.get('gradeLevel', '')); safe_write(ws_front, 'BA66', s2.get('sy', ''))
        safe_write(ws_front, 'BK66', s2.get('sem', '')); safe_write(ws_front, 'G68', s2.get('trackStrand', ''))
        safe_write(ws_front, 'AS68', s2.get('section', ''))

        sub2 = s2.get('subjects', [])
        for i, r in enumerate(range(74, 86)): # 11 rows
            d = sub2[i] if i < len(sub2) else {}
            write_subject_row(ws_front, r, d if d else {})

        # Gen Ave: write our calculated value (template formula is unreliable)
        safe_write(ws_front, 'BD86', to_num(s2.get('genAve', '')))
        safe_write(ws_front, 'F88', s2.get('remarks', ''))
        # --- Signatory (Sem 2) --- PLACEHOLDER CELLS, calibrate as needed
        safe_write(ws_front, 'A92', s2.get('adviserName', ''))
        safe_write(ws_front, 'Y92', s2.get('certName', ''))
        safe_write(ws_front, 'AZ92', s2.get('dateChecked', ''))

        r2 = s2.get('remedial', {})
        if r2:
            safe_write(ws_front, 'S95', r2.get('from', '')); safe_write(ws_front, 'AC95', r2.get('to', ''))
            safe_write(ws_front, 'AL95', r2.get('school', '')); safe_write(ws_front, 'BK95', r2.get('schoolId', ''))
            rem_subs2 = r2.get('subjects', [])
            for i, r in enumerate(range(101, 106)): # 5 remedial rows
                rd = rem_subs2[i] if i < len(rem_subs2) else {}
                write_remedial_row(ws_front, r, rd if rd else {})
            # Remedial teacher/signature — PLACEHOLDER CELLS, calibrate as needed
            safe_write(ws_front, 'J106', r2.get('teacherName', ''))
            safe_write(ws_front, 'AY106', r2.get('signature', ''))

        # ==========================================================
        # GRADE 12 (BACK PAGE)
        # ==========================================================
        
        # --- SEMESTER 3k (G12 - SEM 1) ---
        s3 = data.get('semester3', {})
        safe_write(ws_back, 'E4', s3.get('school', '')); safe_write(ws_back, 'AF4', s3.get('schoolId', ''))
        safe_write(ws_back, 'AS4', s3.get('gradeLevel', '')); safe_write(ws_back, 'BA4', s3.get('sy', ''))
        safe_write(ws_back, 'BK4', s3.get('sem', '')); safe_write(ws_back, 'G5', s3.get('trackStrand', ''))
        safe_write(ws_back, 'AS5', s3.get('section', ''))

        sub3 = s3.get('subjects', [])
        for i, r in enumerate(range(11, 23)):
            d = sub3[i] if i < len(sub3) else {}
            write_subject_row(ws_back, r, d if d else {})

        # Gen Ave: write our calculated value (template formula is unreliable)
        safe_write(ws_back, 'BD23', to_num(s3.get('genAve', '')))
        safe_write(ws_back, 'F25', s3.get('remarks', ''))
        # --- Signatory (Sem 3) --- PLACEHOLDER CELLS, calibrate as needed
        safe_write(ws_back, 'A29', s3.get('adviserName', ''))
        safe_write(ws_back, 'Y29', s3.get('certName', ''))
        safe_write(ws_back, 'AZ29', s3.get('dateChecked', ''))

        r3 = s3.get('remedial', {})
        if r3:
            safe_write(ws_back, 'S32', r3.get('from', '')); safe_write(ws_back, 'AC32', r3.get('to', ''))
            safe_write(ws_back, 'AL32', r3.get('school', '')); safe_write(ws_back, 'BK32', r3.get('schoolId', ''))
            rem_subs3 = r3.get('subjects', [])
            for i, r in enumerate(range(38, 42)): # Q5-Q6 remedial rows
                rd = rem_subs3[i] if i < len(rem_subs3) else {}
                write_remedial_row(ws_back, r, rd if rd else {})
            # Remedial teacher/signature — PLACEHOLDER CELLS, calibrate as needed
            safe_write(ws_back, 'J43', r3.get('teacherName', ''))
            safe_write(ws_back, 'AY43', r3.get('signature', ''))

        # --- SEMESTER 4 (G12 - SEM 2) ---
        s4 = data.get('semester4', {})
        safe_write(ws_back, 'E46', s4.get('school', '')); safe_write(ws_back, 'AF46', s4.get('schoolId', ''))
        safe_write(ws_back, 'AS46', s4.get('gradeLevel', '')); safe_write(ws_back, 'BA46', s4.get('sy', ''))
        safe_write(ws_back, 'BK46', s4.get('sem', '')); safe_write(ws_back, 'G48', s4.get('trackStrand', ''))
        safe_write(ws_back, 'BC48', s4.get('section', ''))

        sub4 = s4.get('subjects', [])
        for i, r in enumerate(range(54, 66)):
            d = sub4[i] if i < len(sub4) else {}
            write_subject_row(ws_back, r, d if d else {})

        # Gen Ave: write our calculated value (template formula is unreliable)
        safe_write(ws_back, 'BD66', to_num(s4.get('genAve', '')))
        safe_write(ws_back, 'F68', s4.get('remarks', ''))
        # --- Signatory (Sem 4) --- PLACEHOLDER CELLS, calibrate as needed
        safe_write(ws_back, 'A72', s4.get('adviserName', ''))
        safe_write(ws_back, 'Y72', s4.get('certName', ''))
        safe_write(ws_back, 'AZ72', s4.get('dateChecked', ''))

        r4 = s4.get('remedial', {})
        if r4:
            safe_write(ws_back, 'S75', r4.get('from', '')); safe_write(ws_back, 'AC75', r4.get('to', ''))
            safe_write(ws_back, 'AL75', r4.get('school', '')); safe_write(ws_back, 'BK75', r4.get('schoolId', ''))
            rem_subs4 = r4.get('subjects', [])
            for i, r in enumerate(range(81, 85)): # 5 remedial rows
                rd = rem_subs4[i] if i < len(rem_subs4) else {}
                write_remedial_row(ws_back, r, rd if rd else {})
            # Remedial teacher/signature — PLACEHOLDER CELLS, calibrate as needed
            safe_write(ws_back, 'I86', r4.get('teacherName', ''))
            safe_write(ws_back, 'AY86', r4.get('signature', ''))

        # --- CERTIFICATION ---
        # Existing
        safe_write(ws_back, 'BI91', cert.get('gradDate', ''))
        safe_write(ws_back, 'BJ90', cert.get('genAve', ''))
        safe_write(ws_back, 'A94', cert.get('schoolHead', ''))
        # Missing fields - PLACEHOLDER CELLS, calibrate as needed
        safe_write(ws_back, 'I90', cert.get('trackStrand', ''))
        safe_write(ws_back, 'I91', cert.get('awards', ''))
        safe_write(ws_back, 'A112', cert.get('remarks', ''))
        safe_write(ws_back, 'T94', cert.get('certDate', ''))
        safe_write(ws_back, 'J114', cert.get('dateIssued', ''))
        
        filename = f"SF10_{str(info.get('lname','Student')).replace(' ','_')}_{int(time.time())}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename); wb.save(output_path)
        
        # If the web client explicitly requests the binary file back automatically
        if return_file:
            # We skip os.startfile and just ship the file out over HTTP
            return send_file(
                output_path,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        # Desktop behavior: open it locally on the bridge machine
        if sys.platform == 'win32': os.startfile(output_path)
        return jsonify({"success": True, "filePath": output_path})
    except Exception as e:
        log_it(f"ERR: {str(e)}"); return jsonify({"success": False, "error": str(e)}), 500

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=PORT, debug=False)
