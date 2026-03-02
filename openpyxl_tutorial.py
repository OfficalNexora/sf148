# Python Tutorial: Modifying Excel with openpyxl

`openpyxl` is the industry standard for Python when you need to **modify existing .xlsx files** without destroying the original formatting, logos, or formulas.

### 1. Installation
Run this in your terminal:
```bash
pip install openpyxl
```

### 2. The Golden Rule of Styles
When you do `cell.value = "New Value"`, `openpyxl` **keeps the existing font, border, and alignment**. It only changes the data inside.

### 3. Basic Example Script
Here is a script that mirrors what we did in Javascript:

```python
import openpyxl
from openpyxl.utils import get_column_letter

def modify_sf10(template_path, output_path, student_data):
    # 1. Load the template (data_only=False keeps your formulas!)
    wb = openpyxl.load_workbook(template_path)
    
    # 2. Select the sheet (by Name)
    ws = wb['FRONT'] 

    # 3. Direct Hardcoding (if you know the coordinates)
    # This is the "Hardcoding" you mentioned
    ws['E23'] = student_data['school']
    ws['AC23'] = student_data['school_id']

    # 4. Search & Write (The "Smart" way)
    def write_near_label(sheet, label_text, value, offset_col=1):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and label_text.upper() in str(cell.value).upper():
                    # Target the cell to the right
                    target_col = cell.column + offset_col
                    sheet.cell(row=cell.row, column=target_col).value = value
                    print(f"Filled {label_text} at {get_column_letter(target_col)}{cell.row}")
                    return True
        return False

    # Example Usage:
    write_near_label(ws, "LAST NAME:", student_data['lname'])
    write_near_label(ws, "FIRST NAME:", student_data['fname'])

    # 5. Saving
    wb.save(output_path)
    print(f"Saved to {output_path}")

# --- Test Case ---
data = {
    'school': 'ACME High',
    'school_id': '123456',
    'lname': 'Dela Cruz',
    'fname': 'Juan'
}

modify_sf10('PLACEHOLDER(ALL).xlsx', 'Output_SF10.xlsx', data)
```

### 4. Handling Merged Cells
If you write to a merged cell (like `SUBJECTS`), you **must** write to the top-left-most cell of that merge. 
- Example: If `A1:C5` is merged, you write to `A1`.
- If you write to `B1`, the data might be hidden or cause an error.

### 5. Why use Python for this?
- **Speed**: Python is very fast at iterating through tens of thousands of cells.
- **Robustness**: `openpyxl` is older and more stable than most JS libraries.
- **Cleanliness**: The syntax is very readable for "hardcoding" your preferences.

**Would you like me to convert the entire Excel Bridge to Python for you, or do you want to keep playing with this library yourself?**
